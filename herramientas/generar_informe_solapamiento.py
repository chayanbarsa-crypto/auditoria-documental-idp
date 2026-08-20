# -*- coding: utf-8 -*-
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from datos_solapamiento import DOCS, TEMAS, CONTRADICCIONES, LITERALES, RECOMENDACION, ESTRUCTURA

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'solapamiento_documentos_modificaciones.xlsx')

NAVY='1F3864'; BLUE='DCE6F1'; GREY='F2F2F2'; RED='F8CBAD'; AMBER='FFE699'; GREEN='E2EFDA'
F_T   = Font(name='Arial', size=14, bold=True, color=NAVY)
F_H2  = Font(name='Arial', size=11, bold=True, color=NAVY)
F_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_B   = Font(name='Arial', size=10)
F_S   = Font(name='Arial', size=9)
F_BD  = Font(name='Arial', size=10, bold=True)
F_IT  = Font(name='Arial', size=9, italic=True)
FILL_H = PatternFill('solid', fgColor=NAVY)
FILL_A = PatternFill('solid', fgColor=GREY)
FILL_R = PatternFill('solid', fgColor=RED)
FILL_M = PatternFill('solid', fgColor=AMBER)
FILL_G = PatternFill('solid', fgColor=GREEN)
FILL_B = PatternFill('solid', fgColor=BLUE)
TH = Side(style='thin', color='BFBFBF'); BOX = Border(left=TH,right=TH,top=TH,bottom=TH)
WRAP = Alignment(wrap_text=True, vertical='top')
CTR  = Alignment(horizontal='center', vertical='center', wrap_text=True)

wb = openpyxl.Workbook()

def header(ws, cols, widths, row=1):
    for j,(h,w) in enumerate(zip(cols,widths), start=1):
        c=ws.cell(row,j,h); c.font=F_HDR; c.fill=FILL_H; c.alignment=CTR; c.border=BOX
        ws.column_dimensions[get_column_letter(j)].width=w

# ================= 00 Resumen =================
ws = wb.active; ws.title='00_Resumen'; ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=3
ws.column_dimensions['B'].width=40
ws.column_dimensions['C'].width=96
r=2
ws.cell(r,2,'Solapamiento entre los 4 documentos de modificaciones').font=F_T; r+=2

ws.cell(r,2,'Documentos analizados').font=F_H2; r+=1
header(ws,['clave','documento','tipo','extension'],[14,62,16,12],r)
r+=1
for k,n,t,e in DOCS:
    for j,v in enumerate([k,n,t,e],start=1):
        c=ws.cell(r,j,v); c.font=F_B; c.alignment=WRAP; c.border=BOX
        if j==3: c.fill = FILL_M if v=='FAQ' else FILL_G
    r+=1
r+=1

ws.cell(r,2,'Hallazgos').font=F_H2; r+=1
H=[('Temas analizados', f'=COUNTA(\'01_Matriz_temas\'!$A$2:$A${len(TEMAS)+1})'),
   ('Temas presentes en 2 o mas documentos', f'=COUNTIF(\'01_Matriz_temas\'!$G$2:$G${len(TEMAS)+1},">=2")'),
   ('Temas presentes en 3 o mas documentos', f'=COUNTIF(\'01_Matriz_temas\'!$G$2:$G${len(TEMAS)+1},">=3")'),
   ('Temas en un solo documento', f'=COUNTIF(\'01_Matriz_temas\'!$G$2:$G${len(TEMAS)+1},1)'),
   ('Contradicciones detectadas', f'=COUNTA(\'03_Contradicciones\'!$A$2:$A${len(CONTRADICCIONES)+1})'),
   ('De ellas, de riesgo ALTA', f'=COUNTIF(\'03_Contradicciones\'!$B$2:$B${len(CONTRADICCIONES)+1},"ALTA")'),
   ('Datos u avisos que un documento tiene y otro OMITE', f'=COUNTIF(\'01_Matriz_temas\'!$B$2:$F${len(TEMAS)+1},"OMITE")')]
RESUMEN_ROWS={}
for lab,f in H:
    RESUMEN_ROWS[lab]=r
    c=ws.cell(r,2,lab); c.font=F_BD; c.border=BOX; c.alignment=WRAP
    c=ws.cell(r,3,f); c.font=F_B; c.border=BOX; c.fill=FILL_A; c.alignment=Alignment(horizontal='left',vertical='center')
    r+=1
r+=1

for titulo, txt in [
 ('El solapamiento es semantico, no copia-pega',
  'Entre los cuatro documentos NO hay una sola coincidencia literal de 9 palabras seguidas. Bajando a 5 palabras '
  'aparecen 20 pasajes entre 000006601 y el FAQs administrativo, y poco mas. Son parafrasis del mismo contenido. '
  'Consecuencia: no se puede detectar el duplicado comparando textos ni por hash, hay que hacerlo por tema.'),
 ('000006601 y el FAQs administrativo son el mismo documento escrito dos veces',
  '11 de los 12 temas administrativos estan en los dos. El unico que no se solapa en todo el conjunto es '
  'Propiedad del contador.'),
 ('Gana el compendio, y es el que tiene la informacion incompleta',
  'El FAQs administrativo se lleva 14 de los primeros puestos de recuperacion medidos. Gana porque cada seccion '
  'lleva un bloque "Ejemplo:" con lenguaje coloquial de agente, que casa con lo que el agente escribe. Pero es el '
  'que omite la restriccion del telefono espanol, la del producto de precio fijo, la opcion de firma en papel y el '
  'metodo de pago por codigo de barras.'),
 ('Cada documento tiene una mitad de la verdad',
  'El aviso de que al desactivar el Monedero Solar se pierde el saldo solo esta en el FAQs. La restriccion del '
  'telefono espanol solo esta en 000006601. Lo que recibe el agente depende de cual gane el ranking ese dia.'),
]:
    ws.cell(r,2,titulo).font=F_BD; ws.cell(r,2).alignment=WRAP
    c=ws.cell(r,3,txt); c.font=F_B; c.alignment=WRAP
    ws.row_dimensions[r].height=14*(1+len(txt)//95)
    r+=1

# ================= 01 Matriz =================
ws = wb.create_sheet('01_Matriz_temas')
header(ws,['Tema','6292 TECNICAS','6601 ADMIN','FAQ-ADM','FAQ-CON','Nota','N docs'],[46,16,14,14,16,58,9])
for i,(tema,a,b,c_,d_,nota) in enumerate(TEMAS):
    r=2+i
    vals=[tema,a,b,c_,d_,nota]
    for j,v in enumerate(vals,start=1):
        cc=ws.cell(r,j, v if v not in ('',None) else None); cc.font=F_B if j in (1,6) else F_S
        cc.alignment=WRAP if j in (1,6) else CTR; cc.border=BOX
        if j in (2,3,4,5):
            if v=='OMITE': cc.fill=FILL_R; cc.font=Font(name='Arial',size=9,bold=True,color='9C0006')
            elif v: cc.fill=FILL_G
    n=ws.cell(r,7, f'=COUNTA($B{r}:$E{r})-COUNTIF($B{r}:$E{r},"OMITE")')
    n.font=F_BD; n.alignment=CTR; n.border=BOX; n.fill=FILL_A
    ws.row_dimensions[r].height=30
ws.freeze_panes='B2'; ws.auto_filter.ref=f'A1:G{len(TEMAS)+1}'

# ================= 02 Literales =================
ws = wb.create_sheet('02_Coincidencias_literales')
ws.cell(1,1,'Pasajes de 5-6 palabras compartidos. No hay ninguna coincidencia de 9 palabras: el solapamiento es de contenido, no de texto.')
ws.cell(1,1).font=F_IT
header(ws,['Par de documentos','Pasaje compartido','Que es'],[24,52,58],2)
for i,(par,frase,que) in enumerate(LITERALES):
    r=3+i
    for j,v in enumerate([par,frase,que],start=1):
        cc=ws.cell(r,j,v); cc.font=F_B if j!=2 else Font(name='Arial',size=10,italic=True)
        cc.alignment=WRAP; cc.border=BOX
    ws.row_dimensions[r].height=26

# ================= 03 Contradicciones =================
ws = wb.create_sheet('03_Contradicciones')
header(ws,['Tema','Riesgo','Documento A','Dice A','Documento B','Dice B','Por que importa'],[30,9,24,44,24,44,52])
for i,t in enumerate(CONTRADICCIONES):
    r=2+i
    for j,v in enumerate(t,start=1):
        cc=ws.cell(r,j,v); cc.font=F_B if j in (1,7) else F_S; cc.alignment=WRAP if j!=2 else CTR; cc.border=BOX
        if j==2: cc.fill = FILL_R if v=='ALTA' else (FILL_M if v=='MEDIA' else FILL_A); cc.font=F_BD
        if j in (4,6): cc.font=Font(name='Arial',size=9,italic=True)
    ws.row_dimensions[r].height=58
ws.freeze_panes='C2'; ws.auto_filter.ref=f'A1:G{len(CONTRADICCIONES)+1}'

# ================= 04 Recomendacion =================
ws = wb.create_sheet('04_Recomendacion')
header(ws,['#','Accion','Documento','Detalle'],[5,14,50,96])
for i,t in enumerate(RECOMENDACION):
    r=2+i
    for j,v in enumerate(t,start=1):
        cc=ws.cell(r,j,v); cc.font=F_BD if j==2 else F_B; cc.alignment=WRAP if j==4 else CTR; cc.border=BOX
        if j==2: cc.fill={'FUSIONAR':FILL_G,'RETIRAR':FILL_R,'PODAR':FILL_M,'DELIMITAR':FILL_M,'UNIFICAR':FILL_B}.get(v,FILL_A)
        if j in (1,3): cc.alignment=WRAP if j==3 else CTR
    ws.row_dimensions[r].height=64

# ================= 05 Estructura FAQ =================
ws = wb.create_sheet('05_Estructura_FAQ')
ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=3
for col,w in zip('BCDE',[30,42,22,74]): ws.column_dimensions[col].width=w
r=2
ws.cell(r,2,'La estructura FAQ, ¿sirve para todo tipo de documento?').font=F_T; r+=2
txt=('No. Sirve muy bien para una cosa concreta y falla en otras dos.\n\n'
     'LO QUE HACE BIEN: los bloques "Ejemplo:" con lenguaje coloquial del agente son lo que hace que un chunk '
     'coincida con lo que el agente escribe. Por eso el FAQs administrativo gana 14 de los primeros puestos. '
     'Ese patron deberia copiarse a TODOS los documentos.\n\n'
     'LO QUE HACE MAL: el formato pregunta-respuesta invita a contestar el caso comun y soltar las excepciones. '
     'Es medible: el FAQs administrativo omite cuatro restricciones que si estan en 000006601 (telefono espanol, '
     'producto de precio fijo, firma en papel, codigo de barras). Y un procedimiento es una secuencia con '
     'precondiciones; troceado en preguntas sueltas pierde el orden.\n\n'
     'CONCLUSION: la FAQ es un COMPLEMENTO, no un sustituto. El patron correcto es documento de procedimiento con '
     'el contenido canonico MAS un bloque de preguntas coloquiales al final. Que es casi lo que ya hace 000006601: '
     'solo le faltan mas preguntas y mas coloquiales.')
c=ws.cell(r,2,txt); c.font=F_B; c.alignment=WRAP
ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
ws.row_dimensions[r].height=200
r+=2
header(ws,['Tipo de documento','Estructura adecuada','¿Formato FAQ?','Por que'],[30,42,22,74],r)
r+=1
for tipo,est,faq,por in ESTRUCTURA:
    for j,v in enumerate([tipo,est,faq,por],start=2):
        cc=ws.cell(r,j,v); cc.font=F_BD if j==4 else F_B; cc.alignment=WRAP if j!=4 else CTR; cc.border=BOX
        if j==4: cc.fill = FILL_G if v.startswith('SI') else FILL_R
    ws.row_dimensions[r].height=52
    r+=1

wb.save(OUT)
print('OK ->', OUT)
print('filas resumen:', RESUMEN_ROWS)
