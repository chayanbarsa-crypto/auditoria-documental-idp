# -*- coding: utf-8 -*-
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from banco_preguntas_lucia import P

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'protocolo_validacion_lucia_v0.xlsx')

# --- paleta ---
NAVY   = '1F3864'
BLUE   = 'DCE6F1'
YELLOW = 'FFF2CC'   # celdas a rellenar
GREY   = 'F2F2F2'
GREEN  = 'E2EFDA'
RED    = 'FCE4E4'

F_TITLE = Font(name='Arial', size=14, bold=True, color=NAVY)
F_H2    = Font(name='Arial', size=11, bold=True, color=NAVY)
F_HDR   = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_BODY  = Font(name='Arial', size=10)
F_SMALL = Font(name='Arial', size=9)
F_BOLD  = Font(name='Arial', size=10, bold=True)

FILL_HDR = PatternFill('solid', fgColor=NAVY)
FILL_IN  = PatternFill('solid', fgColor=YELLOW)
FILL_ALT = PatternFill('solid', fgColor=GREY)
FILL_BOX = PatternFill('solid', fgColor=BLUE)

THIN = Side(style='thin', color='BFBFBF')
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical='top')
CTR  = Alignment(horizontal='center', vertical='center')

N = len(P)                 # 50
R0 = 2                     # primera fila de datos
R1 = R0 + N - 1            # ultima fila de datos  -> 51

wb = openpyxl.Workbook()

# =====================================================================
# 00_Instrucciones
# =====================================================================
ws = wb.active
ws.title = '00_Instrucciones'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 104

r = 2
ws.cell(r, 2, 'Protocolo de validacion de LucIA — banco de 50 preguntas').font = F_TITLE
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1
ws.cell(r, 2, 'Ejecuta las reglas VAL-001, VAL-002 y VAL-003 del checklist, que quedaron FUERA_DE_ALCANCE '
              'en la auditoria auditoria_20260815_135308.xlsx (15 documentos, 34 reglas).').font = F_SMALL
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.row_dimensions[r].height = 26
ws.cell(r, 2).alignment = WRAP

def block(row, title, lines):
    ws.cell(row, 2, title).font = F_H2
    row += 1
    for lab, txt in lines:
        c = ws.cell(row, 2, lab); c.font = F_BOLD; c.alignment = WRAP
        c = ws.cell(row, 3, txt); c.font = F_BODY; c.alignment = WRAP
        ws.row_dimensions[row].height = max(15, 13 * (1 + len(txt) // 100))
        row += 1
    return row + 1

r += 2
r = block(r, 'Para que sirve', [
    ('El problema', 'La auditoria mide la calidad de los PDFs, no las respuestas de LucIA. Sin una medicion de partida '
                    'no se puede demostrar que los documentos corregidos mejoren nada.'),
    ('La solucion', 'Se lanzan las mismas 50 preguntas dos veces: V0 con el corpus actual (ANTES de subir nada) y V1 '
                    'con el corpus corregido. La hoja 04 compara pregunta a pregunta.'),
    ('Regla de oro', 'V0 debe ejecutarse ANTES de subir ningun documento nuevo. Si se sube algo primero, la linea base '
                     'se pierde y el retesting deja de ser comparable.'),
])

r = block(r, 'Como se usa', [
    ('1. Congelar', 'No tocar el corpus. Anotar en la celda C40 la fecha de ejecucion de V0.'),
    ('2. Lanzar V0', 'Copiar cada pregunta de la hoja 02 en LucIA, tal cual, en sesiones independientes '
                     '(sin historial previo, para que una respuesta no contamine la siguiente).'),
    ('3. Puntuar', 'Rellenar SOLO las celdas amarillas de la hoja 02. El veredicto y los puntos se calculan solos.'),
    ('4. Corregir', 'Subir los documentos corregidos al corpus de LucIA.'),
    ('5. Lanzar V1', 'Repetir el proceso en la hoja 03, con las mismas 50 preguntas.'),
    ('6. Comparar', 'Las hojas 04 y 05 muestran la evolucion por pregunta, por categoria y por documento.'),
])

r = block(r, 'Que se puntua en cada pregunta', [
    ('cita_ok  (VAL-002)', 'SI = cita el documento esperado.  PARCIAL = lo cita junto a otros no pertinentes.  '
                           'NO = cita un documento equivocado.  NO_CITA = responde sin citar fuente.'),
    ('exactitud  (VAL-003)', 'LITERAL = coincide con el contenido del documento.  PARCIAL = incompleta pero sin errores.  '
                             'INVENTADA = afirma datos que no estan en la documentacion.  NO_RESPONDE = no contesta.'),
    ('veredicto', 'Se calcula: ACIERTO (1 pto) · PARCIAL (0,5) · FALLO (0) · SIN_RESPUESTA (0) · ALUCINACION (-1).'),
    ('Por que -1', 'Una alucinacion es peor que un silencio: el agente la traslada al cliente como si fuera cierta.'),
])

r = block(r, 'Las seis categorias del banco', [
    ('A_DESAMBIGUACION (14)', 'Productos vecinos que comparten vocabulario. Mide TAX-007/TAX-008/META-006, que fallan '
                              'en los 15 documentos: ninguno tiene bloque "No confundir con".'),
    ('B_SINONIMOS (12)', 'Lenguaje coloquial real de los agentes. Mide TAX-010/META-007, que fallan en los 15.'),
    ('C_TABLAS_PERDIDAS (8)', 'Datos que viven en tablas y capturas que la extraccion de texto perdio. Aqui la respuesta '
                              'correcta suele ser "no dispongo del dato": cualquier cifra concreta es alucinacion.'),
    ('D_TRAMPA (6)', 'Preguntas cuya respuesta correcta es una negativa o un "no consta". Detectan invencion.'),
    ('E_LITERAL (8)', 'Control positivo: datos que si constan literalmente. Un fallo aqui apunta a la recuperacion, '
                      'no a la taxonomia.'),
    ('F_TRAZABILIDAD (2)', 'Normativa de origen. Mide META-009, que solo cumplen 2 de los 15 documentos.'),
])

r = block(r, 'Ejemplo de fila ya cumplimentada (hoja 02 o 03)', [])
r -= 1
ex_hdr = ['id', 'respuesta_de_lucia (E)', 'doc_citado (F)', 'cita_ok (G)', 'exactitud (H)', '-> veredicto', '-> puntos']
ex_val = ['D01',
          'Si, se tramita desde la Oficina del Instalador aportando el certificado de vulnerabilidad.',
          '000001052 Bono Social', 'SI', 'INVENTADA', 'ALUCINACION', '-1']
for i, (h, v) in enumerate(zip(ex_hdr, ex_val)):
    c = ws.cell(r, 2 + (0 if i == 0 else 1))
    if i == 0:
        ws.cell(r, 2, h).font = F_SMALL
        ws.cell(r, 3, v).font = F_BODY
    else:
        ws.cell(r, 2, h).font = F_SMALL
        ws.cell(r, 3, v).font = F_BODY
    ws.cell(r, 2).alignment = WRAP
    ws.cell(r, 3).alignment = WRAP
    r += 1
ws.cell(r, 3, 'Lectura: LucIA cita el documento correcto pero afirma que Fenie Energia puede tramitar el Bono Social. '
              'Cita bien y responde mal — el caso mas peligroso, y el que la auditoria documental no detecta.').font = F_SMALL
ws.cell(r, 3).alignment = WRAP
ws.row_dimensions[r].height = 26
r += 2

r = block(r, 'Leyenda de formato', [
    ('Fondo amarillo', 'Unicas celdas que debes rellenar a mano.'),
    ('Fondo gris', 'Calculado por formula. No escribir encima.'),
    ('Sin fondo', 'Datos de referencia del banco de preguntas.'),
])

ws.cell(39, 2, 'Fecha de ejecucion V0').font = F_BOLD
c = ws.cell(39, 3); c.fill = FILL_IN; c.border = BOX; c.font = F_BODY
ws.cell(40, 2, 'Fecha de ejecucion V1').font = F_BOLD
c = ws.cell(40, 3); c.fill = FILL_IN; c.border = BOX; c.font = F_BODY
ws.cell(41, 2, 'Responsable').font = F_BOLD
c = ws.cell(41, 3); c.fill = FILL_IN; c.border = BOX; c.font = F_BODY

ws.cell(43, 2, 'Origen de los datos').font = F_H2
ws.cell(44, 3, 'Las 50 preguntas se derivan de la evidencia textual citada en auditoria_20260815_135308.xlsx '
               '(hoja Reglas, columna evidencia) y de los metadatos RAG extraidos por documento. '
               'No hay preguntas inventadas al margen del corpus.').font = F_SMALL
ws.cell(44, 3).alignment = WRAP
ws.row_dimensions[44].height = 30

# =====================================================================
# 01_Banco_preguntas
# =====================================================================
wsb = wb.create_sheet('01_Banco_preguntas')
hdr = ['id', 'categoria', 'pregunta', 'doc_esperado_id', 'doc_esperado_nombre',
       'docs_implicados', 'respuesta_correcta_esperada', 'reglas_auditoria', 'fallo_previsto',
       'clave_doc']
widths = [7, 20, 52, 15, 34, 22, 62, 24, 52, 14]
for j, (h, w) in enumerate(zip(hdr, widths), start=1):
    c = wsb.cell(1, j, h); c.font = F_HDR; c.fill = FILL_HDR; c.alignment = CTR; c.border = BOX
    wsb.column_dimensions[get_column_letter(j)].width = w
for i, row in enumerate(P):
    rr = R0 + i
    for j, v in enumerate(row, start=1):
        c = wsb.cell(rr, j, v); c.font = F_BODY; c.alignment = WRAP; c.border = BOX
        if j == 4:
            c.number_format = '@'
    # clave_doc: identificador NO numerico. COUNTIF/SUMIFS coercionan a numero
    # cualquier criterio que parezca un numero ("000001052" -> 1052), lo que no
    # casaria con el texto del rango. El prefijo "D" evita esa coercion.
    c = wsb.cell(rr, 10, 'D' + row[3])
    c.font = F_SMALL; c.alignment = WRAP; c.border = BOX; c.number_format = '@'
    wsb.row_dimensions[rr].height = 46
wsb.cell(1, 10).comment = None
wsb.freeze_panes = 'C2'
wsb.auto_filter.ref = f'A1:J{R1}'

# =====================================================================
# 02 / 03  hojas de ejecucion
# =====================================================================
EX_HDR = ['id', 'categoria', 'pregunta', 'doc_esperado',
          'respuesta_de_lucia', 'doc_citado_por_lucia', 'cita_ok', 'exactitud',
          'veredicto', 'puntos', 'observaciones']
EX_W   = [7, 20, 52, 34, 60, 26, 12, 14, 16, 8, 40]

def build_exec(name, etiqueta):
    w = wb.create_sheet(name)
    t = w.cell(1, 1, f'EJECUCION {etiqueta}')
    for j, (h, wd) in enumerate(zip(EX_HDR, EX_W), start=1):
        c = w.cell(1, j, h); c.font = F_HDR; c.fill = FILL_HDR; c.alignment = CTR; c.border = BOX
        w.column_dimensions[get_column_letter(j)].width = wd
    for i in range(N):
        rr = R0 + i
        w.cell(rr, 1, f"='01_Banco_preguntas'!A{rr}")
        w.cell(rr, 2, f"='01_Banco_preguntas'!B{rr}")
        w.cell(rr, 3, f"='01_Banco_preguntas'!C{rr}")
        w.cell(rr, 4, f"='01_Banco_preguntas'!D{rr}&\" \"&'01_Banco_preguntas'!E{rr}")
        # entradas manuales
        for j in (5, 6, 7, 8, 11):
            c = w.cell(rr, j); c.fill = FILL_IN
        # veredicto
        w.cell(rr, 9,
               f'=IF(OR($G{rr}="",$H{rr}=""),"",'
               f'IF($H{rr}="INVENTADA","ALUCINACION",'
               f'IF($H{rr}="NO_RESPONDE","SIN_RESPUESTA",'
               f'IF(AND($G{rr}="SI",$H{rr}="LITERAL"),"ACIERTO",'
               f'IF(OR($G{rr}="NO",$G{rr}="NO_CITA"),"FALLO","PARCIAL")))))')
        # puntos
        w.cell(rr, 10,
               f'=IF($I{rr}="","",'
               f'IF($I{rr}="ACIERTO",1,'
               f'IF($I{rr}="PARCIAL",0.5,'
               f'IF($I{rr}="ALUCINACION",-1,0))))')
        for j in (9, 10):
            w.cell(rr, j).fill = FILL_ALT
        for j in range(1, 12):
            c = w.cell(rr, j); c.font = F_BODY; c.alignment = WRAP; c.border = BOX
        w.cell(rr, 7).alignment = CTR
        w.cell(rr, 8).alignment = CTR
        w.cell(rr, 9).alignment = CTR
        w.cell(rr, 10).alignment = CTR
        w.cell(rr, 10).number_format = '0.0'
        w.row_dimensions[rr].height = 44
    dv1 = DataValidation(type='list', formula1='"SI,PARCIAL,NO,NO_CITA"', allow_blank=True, showDropDown=False)
    dv2 = DataValidation(type='list', formula1='"LITERAL,PARCIAL,INVENTADA,NO_RESPONDE"', allow_blank=True, showDropDown=False)
    w.add_data_validation(dv1); w.add_data_validation(dv2)
    dv1.add(f'G{R0}:G{R1}'); dv2.add(f'H{R0}:H{R1}')
    w.freeze_panes = 'E2'
    w.auto_filter.ref = f'A1:K{R1}'
    return w

build_exec('02_Ejecucion_V0', 'V0 — LINEA BASE (corpus actual, antes de subir correcciones)')
build_exec('03_Ejecucion_V1', 'V1 — RETESTING (corpus corregido)')

# =====================================================================
# 04_Comparativa
# =====================================================================
wc = wb.create_sheet('04_Comparativa')
CH = ['id', 'categoria', 'pregunta', 'doc_esperado_id',
      'veredicto_V0', 'puntos_V0', 'veredicto_V1', 'puntos_V1', 'delta', 'evolucion']
CW = [7, 20, 52, 15, 16, 10, 16, 10, 9, 14]
for j, (h, wd) in enumerate(zip(CH, CW), start=1):
    c = wc.cell(1, j, h); c.font = F_HDR; c.fill = FILL_HDR; c.alignment = CTR; c.border = BOX
    wc.column_dimensions[get_column_letter(j)].width = wd
for i in range(N):
    rr = R0 + i
    wc.cell(rr, 1, f"='01_Banco_preguntas'!A{rr}")
    wc.cell(rr, 2, f"='01_Banco_preguntas'!B{rr}")
    wc.cell(rr, 3, f"='01_Banco_preguntas'!C{rr}")
    wc.cell(rr, 4, f"='01_Banco_preguntas'!D{rr}")
    wc.cell(rr, 5, f"=IF('02_Ejecucion_V0'!I{rr}=\"\",\"\",'02_Ejecucion_V0'!I{rr})")
    wc.cell(rr, 6, f"=IF('02_Ejecucion_V0'!J{rr}=\"\",\"\",'02_Ejecucion_V0'!J{rr})")
    wc.cell(rr, 7, f"=IF('03_Ejecucion_V1'!I{rr}=\"\",\"\",'03_Ejecucion_V1'!I{rr})")
    wc.cell(rr, 8, f"=IF('03_Ejecucion_V1'!J{rr}=\"\",\"\",'03_Ejecucion_V1'!J{rr})")
    wc.cell(rr, 9, f'=IF(OR($F{rr}="",$H{rr}=""),"",$H{rr}-$F{rr})')
    wc.cell(rr, 10, f'=IF($I{rr}="","",IF($I{rr}>0,"MEJORA",IF($I{rr}<0,"REGRESION","IGUAL")))')
    for j in range(1, 11):
        c = wc.cell(rr, j); c.font = F_BODY; c.alignment = WRAP; c.border = BOX
        if j >= 5:
            c.fill = FILL_ALT; c.alignment = CTR
    for j in (6, 8, 9):
        wc.cell(rr, j).number_format = '0.0'
    wc.row_dimensions[rr].height = 30
wc.freeze_panes = 'E2'
wc.auto_filter.ref = f'A1:J{R1}'

# =====================================================================
# 05_Panel
# =====================================================================
wp = wb.create_sheet('05_Panel')
wp.sheet_view.showGridLines = False
for col, wd in zip('ABCDEF', [3, 34, 14, 14, 14, 46]):
    wp.column_dimensions[col].width = wd

wp.cell(2, 2, 'Panel de resultados — LucIA V0 vs V1').font = F_TITLE

V0I = f"'02_Ejecucion_V0'!$I${R0}:$I${R1}"
V0J = f"'02_Ejecucion_V0'!$J${R0}:$J${R1}"
V1I = f"'03_Ejecucion_V1'!$I${R0}:$I${R1}"
V1J = f"'03_Ejecucion_V1'!$J${R0}:$J${R1}"
BCAT = f"'01_Banco_preguntas'!$B${R0}:$B${R1}"
BDOC = f"'01_Banco_preguntas'!$J${R0}:$J${R1}"   # clave_doc, no numerica

def hdr_row(r, cols):
    for j, t in enumerate(cols, start=2):
        c = wp.cell(r, j, t); c.font = F_HDR; c.fill = FILL_HDR; c.alignment = CTR; c.border = BOX

# --- bloque 1: recuento por veredicto ---
r = 4
wp.cell(r, 2, '1 · Recuento por veredicto').font = F_H2
r += 1
hdr_row(r, ['veredicto', 'V0', 'V1', 'delta', 'lectura'])
lect = {
 'ACIERTO':      'Cita el documento correcto y responde con su contenido.',
 'PARCIAL':      'Se acerca: cita parcial o respuesta incompleta pero sin errores.',
 'FALLO':        'Cita el documento equivocado o no cita ninguno.',
 'SIN_RESPUESTA':'No contesta. Preferible a inventar.',
 'ALUCINACION':  'Afirma datos que no estan en la documentacion. El riesgo real.',
}
r += 1
first_v = r
for v, txt in lect.items():
    wp.cell(r, 2, v).font = F_BODY
    wp.cell(r, 3, f'=COUNTIF({V0I},$B{r})')
    wp.cell(r, 4, f'=COUNTIF({V1I},$B{r})')
    wp.cell(r, 5, f'=$D{r}-$C{r}')
    wp.cell(r, 6, txt).font = F_SMALL
    for j in range(2, 7):
        c = wp.cell(r, j); c.border = BOX; c.alignment = WRAP if j == 6 else CTR
        if j == 2: c.alignment = Alignment(vertical='center')
        if j in (3, 4, 5): c.fill = FILL_ALT; c.font = F_BODY
    r += 1
last_v = r - 1
wp.cell(r, 2, 'Preguntas evaluadas').font = F_BOLD
wp.cell(r, 3, f'=SUM($C${first_v}:$C${last_v})')
wp.cell(r, 4, f'=SUM($D${first_v}:$D${last_v})')
wp.cell(r, 5, f'=$D{r}-$C{r}')
for j in range(2, 6):
    c = wp.cell(r, j); c.border = BOX; c.fill = FILL_BOX; c.font = F_BOLD; c.alignment = CTR
wp.cell(r, 2).alignment = Alignment(vertical='center')
tot_row = r

# --- bloque 2: KPIs ---
r += 3
wp.cell(r, 2, '2 · Indicadores').font = F_H2
r += 1
hdr_row(r, ['indicador', 'V0', 'V1', 'delta', 'objetivo'])
r += 1
kpis = [
 ('Tasa de acierto',
  f'=IF($C${tot_row}=0,"",COUNTIF({V0I},"ACIERTO")/$C${tot_row})',
  f'=IF($D${tot_row}=0,"",COUNTIF({V1I},"ACIERTO")/$D${tot_row})',
  '0.0%', 'Subir. Es el KPI principal del retesting.'),
 ('Tasa de alucinacion',
  f'=IF($C${tot_row}=0,"",COUNTIF({V0I},"ALUCINACION")/$C${tot_row})',
  f'=IF($D${tot_row}=0,"",COUNTIF({V1I},"ALUCINACION")/$D${tot_row})',
  '0.0%', 'Bajar a 0. Criterio de bloqueo para pasar a produccion.'),
 ('Tasa de cita correcta (VAL-002)',
  f'=IF($C${tot_row}=0,"",COUNTIF(\'02_Ejecucion_V0\'!$G${R0}:$G${R1},"SI")/$C${tot_row})',
  f'=IF($D${tot_row}=0,"",COUNTIF(\'03_Ejecucion_V1\'!$G${R0}:$G${R1},"SI")/$D${tot_row})',
  '0.0%', 'Mide si LucIA recupera el documento correcto.'),
 ('Puntuacion total',
  f'=IF($C${tot_row}=0,"",SUM({V0J}))',
  f'=IF($D${tot_row}=0,"",SUM({V1J}))',
  '0.0', f'Maximo posible: {N},0'),
 ('Puntuacion media por pregunta',
  f'=IF($C${tot_row}=0,"",SUM({V0J})/$C${tot_row})',
  f'=IF($D${tot_row}=0,"",SUM({V1J})/$D${tot_row})',
  '0.00', 'Comparable aunque V0 y V1 tengan distinto numero de preguntas evaluadas.'),
]
for lab, f0, f1, fmt, obj in kpis:
    wp.cell(r, 2, lab).font = F_BODY
    wp.cell(r, 3, f0); wp.cell(r, 4, f1)
    wp.cell(r, 5, f'=IF(OR($C{r}="",$D{r}=""),"",$D{r}-$C{r})')
    wp.cell(r, 6, obj).font = F_SMALL
    for j in (3, 4, 5):
        c = wp.cell(r, j); c.number_format = fmt; c.fill = FILL_ALT; c.font = F_BODY
    for j in range(2, 7):
        c = wp.cell(r, j); c.border = BOX
        c.alignment = WRAP if j == 6 else (CTR if j > 2 else Alignment(vertical='center'))
    r += 1

# --- bloque 3: por categoria ---
r += 2
wp.cell(r, 2, '3 · Puntuacion por categoria').font = F_H2
r += 1
hdr_row(r, ['categoria', 'V0', 'V1', 'delta', 'que mide'])
r += 1
cats = [
 ('A_DESAMBIGUACION', 'Productos vecinos. TAX-007/TAX-008/META-006 fallan en 15/15 documentos.'),
 ('B_SINONIMOS',      'Lenguaje coloquial del agente. TAX-010/META-007 fallan en 15/15.'),
 ('C_TABLAS_PERDIDAS','Datos en tablas y capturas que la extraccion perdio. 17 reglas SIN_EVIDENCIA.'),
 ('D_TRAMPA',         'Respuesta correcta = negativa o "no consta". Detecta invencion.'),
 ('E_LITERAL',        'Control positivo: el dato si consta. Un fallo apunta a la recuperacion.'),
 ('F_TRAZABILIDAD',   'Normativa de origen. META-009 solo lo cumplen 2 de 15 documentos.'),
]
for cat, txt in cats:
    wp.cell(r, 2, cat).font = F_BODY
    wp.cell(r, 3, f'=SUMIFS({V0J},{BCAT},$B{r})')
    wp.cell(r, 4, f'=SUMIFS({V1J},{BCAT},$B{r})')
    wp.cell(r, 5, f'=$D{r}-$C{r}')
    wp.cell(r, 6, txt).font = F_SMALL
    for j in (3, 4, 5):
        c = wp.cell(r, j); c.number_format = '0.0'; c.fill = FILL_ALT; c.font = F_BODY
    for j in range(2, 7):
        c = wp.cell(r, j); c.border = BOX
        c.alignment = WRAP if j == 6 else (CTR if j > 2 else Alignment(vertical='center'))
    r += 1

# --- bloque 4: por documento ---
docs = []
seen = set()
for row in P:
    if row[3] not in seen:
        seen.add(row[3]); docs.append((row[3], row[4]))
docs.sort()

r += 2
wp.cell(r, 2, '4 · Cobertura y puntuacion por documento').font = F_H2
r += 1
hdr_row(r, ['document_id', 'preguntas', 'V0', 'V1', 'documento'])
r += 1
for did, dname in docs:
    c = wp.cell(r, 2, did); c.font = F_BODY; c.number_format = '@'
    wp.cell(r, 3, f'=COUNTIF({BDOC},"D"&$B{r})')
    wp.cell(r, 4, f'=SUMIFS({V0J},{BDOC},"D"&$B{r})')
    wp.cell(r, 5, f'=SUMIFS({V1J},{BDOC},"D"&$B{r})')
    wp.cell(r, 6, dname).font = F_SMALL
    for j in (3, 4, 5):
        cc = wp.cell(r, j); cc.number_format = '0.0'; cc.fill = FILL_ALT; cc.font = F_BODY
    for j in range(2, 7):
        cc = wp.cell(r, j); cc.border = BOX
        cc.alignment = WRAP if j == 6 else (CTR if j > 2 else Alignment(vertical='center'))
    r += 1
wp.cell(r, 2, 'TOTAL').font = F_BOLD
wp.cell(r, 3, f'=SUM($C${r-len(docs)}:$C${r-1})')
wp.cell(r, 4, f'=SUM($D${r-len(docs)}:$D${r-1})')
wp.cell(r, 5, f'=SUM($E${r-len(docs)}:$E${r-1})')
for j in range(2, 6):
    c = wp.cell(r, j); c.border = BOX; c.fill = FILL_BOX; c.font = F_BOLD; c.alignment = CTR
wp.cell(r, 2).alignment = Alignment(vertical='center')
for j in (4, 5):
    wp.cell(r, j).number_format = '0.0'

wb.save(OUT)
print('OK ->', OUT)
